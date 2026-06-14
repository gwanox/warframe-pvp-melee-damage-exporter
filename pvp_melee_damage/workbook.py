"""Excel workbook writer."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from .constants import (
    COMBO_CONTEXT_HEADERS,
    COMBO_CONTEXT_SOURCE,
    ENUM_REFERENCE_SOURCE,
    PIVOT_HEADERS,
    PVP_STANCE_NAME_SOURCE,
    SLAM_HEADERS,
    TOTAL_HEADERS,
    WARNING_HEADERS,
)
from .models import StanceInfo
from .rows import build_combo_context_rows, build_compact_tables
from .utils import rows_to_matrix, sorted_warnings


def write_workbook(
    output_path: Path,
    pivot_rows: list[dict[str, Any]],
    total_rows: list[dict[str, Any]],
    slam_rows: list[dict[str, Any]],
    warnings: list[dict[str, str]],
    misc_stances: list[StanceInfo],
    command: str,
    layout: str = "thematic",
    include_non_tonfa_air_right: bool = False,
    include_unverified_pvp_contexts: bool = False,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise SystemExit(
            "Missing dependency: openpyxl. Install it with:\n"
            "  py -m pip install -r Melee_attack_dmg_analysis/requirements.txt"
        ) from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    warning_rows = sorted_warnings(warnings)

    def add_sheet(
        name: str, headers: list[str], data_rows: list[list[Any]], table_name: str | None = None
    ) -> None:
        ws = workbook.create_sheet(name)
        ws.append(headers)
        for data_row in data_rows:
            ws.append(data_row)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        if table_name and data_rows:
            end_col = get_column_letter(len(headers))
            end_row = len(data_rows) + 1
            table = Table(displayName=table_name, ref=f"A1:{end_col}{end_row}")
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

        for column_index, header in enumerate(headers, start=1):
            width = min(max(len(header) + 2, 12), 52)
            for row_index in range(2, min(ws.max_row, 80) + 1):
                width = min(
                    max(
                        width, len(str(ws.cell(row=row_index, column=column_index).value or "")) + 2
                    ),
                    52,
                )
            ws.column_dimensions[get_column_letter(column_index)].width = width

    if layout == "raw":
        add_sheet("Raw Data", PIVOT_HEADERS, rows_to_matrix(PIVOT_HEADERS, pivot_rows))
        add_sheet("Slams", SLAM_HEADERS, rows_to_matrix(SLAM_HEADERS, slam_rows))
        add_sheet("Warnings", WARNING_HEADERS, rows_to_matrix(WARNING_HEADERS, warning_rows))
    elif layout == "full":
        add_sheet(
            "Hit Damage Database",
            PIVOT_HEADERS,
            rows_to_matrix(PIVOT_HEADERS, pivot_rows),
            "HitDamageDatabase",
        )
        add_sheet(
            "Combo Damage Database",
            TOTAL_HEADERS,
            rows_to_matrix(TOTAL_HEADERS, total_rows),
            "ComboDamageDatabase",
        )
        add_sheet("Slams", SLAM_HEADERS, rows_to_matrix(SLAM_HEADERS, slam_rows), "SlamsTable")
        add_sheet(
            "Combo Context",
            COMBO_CONTEXT_HEADERS,
            rows_to_matrix(
                COMBO_CONTEXT_HEADERS,
                build_combo_context_rows(
                    pivot_rows,
                    include_non_tonfa_air_right,
                    include_unverified_pvp_contexts,
                ),
            ),
            "ComboContext",
        )
        add_sheet(
            "Warnings",
            WARNING_HEADERS,
            rows_to_matrix(WARNING_HEADERS, warning_rows),
            "WarningsTable",
        )
    else:
        compact_tables = build_compact_tables(pivot_rows, slam_rows, misc_stances)
        headers, data_rows = compact_tables["Hit Damage Database"]
        add_sheet("Hit Damage Database", headers, data_rows)
        add_sheet("Combo Damage Database", TOTAL_HEADERS, rows_to_matrix(TOTAL_HEADERS, total_rows))
        for name in ("Weapons", "Stances", "Combos", "Attacks"):
            headers, data_rows = compact_tables[name]
            add_sheet(name, headers, data_rows)
        add_sheet("Slams", SLAM_HEADERS, rows_to_matrix(SLAM_HEADERS, slam_rows))
        add_sheet(
            "Combo Context",
            COMBO_CONTEXT_HEADERS,
            rows_to_matrix(
                COMBO_CONTEXT_HEADERS,
                build_combo_context_rows(
                    pivot_rows,
                    include_non_tonfa_air_right,
                    include_unverified_pvp_contexts,
                ),
            ),
        )
        add_sheet("Warnings", WARNING_HEADERS, rows_to_matrix(WARNING_HEADERS, warning_rows))

    readme = workbook.create_sheet("Readme")
    readme_rows = [
        ("Generated", datetime.now().isoformat(timespec="seconds")),
        ("Command", command),
        ("Layout", layout),
        (
            "Non-Tonfa CC_AIR_RIGHT",
            "included" if include_non_tonfa_air_right else "excluded",
        ),
        (
            "Unverified PvP contexts",
            "included" if include_unverified_pvp_contexts else "excluded",
        ),
        ("Hit rows", len(pivot_rows)),
        ("Combo damage rows", len(total_rows)),
        ("Slam rows", len(slam_rows)),
        ("Warnings rows", len(warnings)),
        ("", ""),
        ("Section", "Methodology"),
        (
            "Methodology",
            "Resolve JSON package inheritance from base package to derived package using deep dictionary merge; lists and scalars are replaced.",
        ),
        (
            "Methodology",
            "Discover PvP melee weapons by scanning Lotus/Weapons/**/*.json for melee sweep behavior markers, then requiring effective AvailableOnPvp=1 unless force-included.",
        ),
        (
            "Methodology",
            "Read base damage and PvP multiplier from the melee sweep behavior's MeleeImpactBehavior AttackData and PvpDamageMultiplier.",
        ),
        (
            "Methodology",
            "Read weapon attack speed from the same melee behavior's state fireRate and divide by 60 to convert attacks per minute to attacks per second.",
        ),
        (
            "Methodology",
            "Read InitialHitCounter, BaseHitMultipler, HitReqNextTierOperator, and HitReqNextTierOperationType from the melee impact behavior to derive the starting combo multiplier.",
        ),
        (
            "Methodology",
            "Apply the starting combo multiplier before final rounding to CC_GROUND_HEAVY and to PvpSlams entries with CanUseComboMultiplier=1; source attack/slam attenuation columns remain unchanged.",
        ),
        (
            "Methodology",
            "Normalize weapon Impact, Puncture, Slash, and elemental damage from AttackData; old-format component fractions are multiplied by Amount, while new-format components are already damage values.",
        ),
        (
            "Methodology",
            "Create Scindo/Scindo Prime Manticore and Fragor/Fragor Prime Brokk variants from the local legacy skin Upgrades. Prime compatibility is an explicit mapping because each skin package names only its base weapon.",
        ),
        (
            "Methodology",
            "Match PvP stances by inherited melee tree package or compatibility tags. stance_equipped=no comes from the base melee tree; stance_equipped=yes comes from the stance-derived tree; identical rows collapse to any.",
        ),
        (
            "Methodology",
            "stance_id identifies the compatible comparison stance even on no/any rows; it does not by itself mean the stance is installed.",
        ),
        (
            "Methodology",
            "Use the effective EquippedAttackSets map for current PvP attacks. stance_equipped=no resolves that map from the base melee tree, while stance_equipped=yes resolves the stance-derived override; UnequippedAttackSets is retained in the files as legacy quick-melee routing and is not exported.",
        ),
        (
            "Methodology",
            "Expand ContextFallbacks only within the same attack-set map and annotate the resulting rows. A context explicitly present in the map, including an empty value, blocks its fallback.",
        ),
        (
            "Methodology",
            "Treat the context key in EquippedAttackSets as the runtime trigger. Attack-set ComboContext metadata is not used to relabel rows because inherited and reused sets frequently retain a different internal context.",
        ),
        (
            "Methodology",
            "Name PvP stances from the actual Conclave stance list; misc stance packages are excluded from damage rows and summarized once in Stances.",
        ),
        (
            "Methodology",
            "Calculate each hit as round_half_up(base_damage * pvp_multiplier * effective_attack_atten * quant_multiplier * eligible_initial_combo_multiplier).",
        ),
        (
            "Methodology",
            "Read PvP aerial slam and heavy slam rows from each weapon's effective PvpSlams entries for MeleeSlam and HeavySlam.",
        ),
        (
            "Methodology",
            "Calculate each PvP slam as round_half_up(base_damage * pvp_multiplier * BaseDamageAttenuation * quant_multiplier * eligible_initial_combo_multiplier); radius and falloff come from the same PvpSlams entry.",
        ),
        (
            "Methodology",
            "Append core PvP slams to Hit Damage Database and Combo Damage Database as one-hit rows using the existing combo/attack columns.",
        ),
        (
            "Methodology",
            "PvpSlams may include a MeleeAttack reference, but it is not exported as a column because the concrete PvP damage/radius fields come from the PvpSlams entry.",
        ),
        (
            "Data note",
            "Dark Split-Sword PvpSlams omits HeavySlam for both modes, so heavy-slam cells remain blank and PvE Slams.HeavySlam is not used as a fallback. In-game PvP testing confirms the dual-mode heavy slam glitches.",
        ),
        (
            "Data note",
            "Brokk's -1 second combo-duration modifier is retained in Weapons.note and Warnings but does not change hit, combo-damage, or slam calculations.",
        ),
        (
            "Data note",
            "Synoid Heliocor and Furax Wraith have 20 initial combo, while Fragor Prime has 30; all derive a 2x starting heavy-attack multiplier. Fragor Prime's value is inherited by its Brokk Skin variant.",
        ),
        (
            "Data note",
            "Initial combo is a PvE-oriented mechanic that remains active in PvP. The exporter treats this as a known Conclave carryover and includes it in eligible heavy attack and heavy slam damage.",
        ),
        (
            "Methodology",
            "Effective attack attenuation prefers per-swing PvP override, then attack PvP BaseDamageAtten, then regular attack BaseDamageAtten, then 1.",
        ),
        (
            "Methodology",
            "Physical-only weapons use IPS quantization from rounded 32nds; elemental/non-physical rows use quant multiplier 1.",
        ),
        (
            "Methodology",
            "Combo damage is the ordered sum of exported hit damages; damage_instances shows that ordered hit list before total_damage.",
        ),
        (
            "Methodology",
            "Export resolvable contexts from effective EquippedAttackSets except known legacy charge aliases stored under CC_SLIDING_PVP. CC_ATTACK_BLOCKED, CC_DOWNED_ENEMY, CC_GUARD_BROKEN, and CC_PARRY_HEAVY are excluded by default because current PvP testing could not trigger them; --include-unverified-pvp-contexts restores them for source-data auditing.",
        ),
        (
            "Data note",
            "Most heavy attacks are inherited unchanged from the base melee tree and therefore use stance_equipped=any. Biting Piranha explicitly overrides the Dual Dagger heavy attack, so its no/yes rows remain distinct.",
        ),
        (
            "Data note",
            "Combos.weapon_scope distinguishes category-wide mappings from rows produced by special weapons or modes; weapon_names identifies the exact scope.",
        ),
        (
            "Data note",
            "Current slide attacks resolve through EquippedAttackSets.CC_SLIDING. Skana resolves to one 68-damage hit with or without Rising Steel. Star Divide overrides the Tonfa slide with PVPTonfaSlideAEquipped; in-game testing verifies two 1x hits, including 59 + 59 on Kronen, while the stanceless Tonfa slide remains one 6x hit (354 on Kronen).",
        ),
        (
            "Data note",
            "Hit counts are inferred from effective PerSwingOverrides. Empty derived attack packages may be engine-side runtime markers whose true hit behavior is not recoverable from inherited JSON alone; unvalidated cases are listed in Warnings.",
        ),
        (
            "Data note",
            (
                "Direct AerialAttacks button-routing refs are outside the attack-set context model and are not exported as combo rows; CC_AIR_RIGHT attack sets are exported for all categories by request."
                if include_non_tonfa_air_right
                else "Direct AerialAttacks button-routing refs are outside the attack-set context model and are not exported as combo rows; CC_AIR_RIGHT is retained for Tonfas but omitted for other categories because their A/B aerial sets currently produce identical PvP damage."
            ),
        ),
        (
            "Data note",
            "Damage rows model melee sweep impacts only. Projectile throws use separate fire behaviors and are not calculated as blade hits, including glaive throws and the Sigma and Octantis aerial shield throw.",
        ),
        ("Quantization reference", "https://wiki.warframe.com/w/Damage/Calculation"),
        ("Initial combo reference", "https://wiki.warframe.com/w/Melee#Combo_Counter"),
        ("Fragor Prime reference", "https://wiki.warframe.com/w/Fragor_Prime"),
        ("Combo context reference", COMBO_CONTEXT_SOURCE),
        ("Enum reference", ENUM_REFERENCE_SOURCE),
        ("Stance name reference", PVP_STANCE_NAME_SOURCE),
        ("", ""),
        ("Section", "Assumptions"),
        (
            "Assumption",
            "Thematic layout writes human-readable damage database sheets plus raw reference sheets and one combo-damage sum sheet.",
        ),
        ("Assumption", "Raw layout writes one denormalized data sheet."),
        ("Assumption", "Full layout keeps the older denormalized Excel table sheets."),
        (
            "Assumption",
            "Weapon source scans Lotus/Weapons/**/*.json for melee sweep behavior markers.",
        ),
        (
            "Assumption",
            "Weapon category labels are canonicalized against https://warframe.fandom.com/wiki/Category:Melee_Weapon_Type.",
        ),
        (
            "Assumption",
            "Known not-in-game and AI-like weapon files are excluded and listed as ignored warnings.",
        ),
        (
            "Assumption",
            "Identical base-tree/stance-derived rows use stance_equipped = any; weapon-level slam rows also use any because slams are independent of stance attack sets.",
        ),
        (
            "Assumption",
            "Stances sheet omits stance_equipped, keeps actual stance rows, and adds one summary row per excluded misc stance package.",
        ),
        (
            "Assumption",
            "Paracesis and Broken War no-stance states use embedded stance trees from local metadata.",
        ),
        (
            "Assumption",
            "Embedded no-stance rows use unique category labels such as Paracesis (stanceless) and Broken War (stanceless).",
        ),
        (
            "Assumption",
            "Weapons.note flags embedded stances, forced modes, legacy skin modifiers, and known unique behavior such as Sigma and Octantis aerial shield throw.",
        ),
        (
            "Assumption",
            "Slams exports core aerial PvP slams only: MeleeSlam and HeavySlam. Non-core slam events with nonzero AttackData.Amount are noted in Warnings for later audit.",
        ),
        ("Assumption", "Rows are per hit/swing; ImpulseAtten is ignored for HP damage."),
    ]
    for row in readme_rows:
        readme.append(row)

    section_fill = PatternFill("solid", fgColor="F3F4F6")
    readme.column_dimensions["A"].width = 32
    readme.column_dimensions["B"].width = 175
    for row_index, row in enumerate(readme_rows, start=1):
        label_cell = readme.cell(row=row_index, column=1)
        value_cell = readme.cell(row=row_index, column=2)
        label_cell.alignment = Alignment(vertical="center")
        value_cell.alignment = Alignment(vertical="center", wrap_text=False)

        if row[0] and row[0] not in {"Methodology", "Assumption", "Tip"}:
            label_cell.font = Font(bold=True)

        if row[0] == "Section":
            label_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            label_cell.font = Font(bold=True)
            value_cell.font = Font(bold=True)
            label_cell.fill = section_fill
            value_cell.fill = section_fill
        elif row[0] == "":
            readme.row_dimensions[row_index].height = 18

    workbook.save(output_path)
